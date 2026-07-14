from app.parsers.base import BaseParser
import fitz
import io
import docx

class PDFParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        text = ""
        try:
            with fitz.open(stream=content, filetype="pdf") as doc:
                for page in doc:
                    text += page.get_text()
        except Exception as e:
            raise ValueError(f"Failed to parse PDF: {str(e)}")
        return text

class DOCXParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        try:
            doc = docx.Document(io.BytesIO(content))
            return "\n".join([para.text for para in doc.paragraphs])
        except Exception as e:
            raise ValueError(f"Failed to parse DOCX: {str(e)}")

class LegacyDOCParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        raise ValueError("Legacy DOC format is not supported. Please convert to DOCX or PDF.")

class TXTParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        return content.decode("utf-8", errors="replace")

class MarkdownParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        return content.decode("utf-8", errors="replace")

class GoogleDocsParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        # Google Docs are exported as text/plain
        return content.decode("utf-8", errors="replace")

import csv
class CSVParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        try:
            # Try decoding as utf-8, fallback to latin-1
            text = content.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            return "\n".join([", ".join(row) for row in reader])
        except Exception as e:
            raise ValueError(f"Failed to parse CSV: {str(e)}")

import openpyxl
class XLSXParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            output = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                output.append(f"--- Sheet: {sheet_name} ---")
                for row in sheet.iter_rows(values_only=True):
                    row_str = ", ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_str.strip(", "):
                        output.append(row_str)
            return "\n".join(output)
        except Exception as e:
            raise ValueError(f"Failed to parse XLSX: {str(e)}")

class GoogleSheetsParser(BaseParser):
    def parse(self, content: bytes, mime_type: str) -> str:
        # Google Sheets are usually exported as text/csv or xlsx
        # If text/csv, we can reuse CSVParser logic
        parser = CSVParser()
        return parser.parse(content, mime_type)


class ParserFactory:
    """
    Factory to retrieve the correct parser based on MIME type.
    """
    _parsers = {
        "application/pdf": PDFParser(),
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": DOCXParser(),
        "application/msword": LegacyDOCParser(),
        "text/plain": TXTParser(),
        "text/csv": CSVParser(),
        "text/markdown": MarkdownParser(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": XLSXParser(),
        "application/vnd.google-apps.document": GoogleDocsParser(), # Note: Actually fetched as text/plain by the connector export
        "application/vnd.google-apps.spreadsheet": GoogleSheetsParser(), # Actually exported as text/csv
    }

    _extension_map = {
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".doc": "application/msword",
        ".txt": "text/plain",
        ".csv": "text/csv",
        ".md": "text/markdown",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    @classmethod
    def get_parser(cls, mime_type: str, filename: str | None = None) -> BaseParser:
        resolved_mime = mime_type
        
        # If generic/unknown or unsupported MIME type, try resolving via file extension
        if mime_type in ["application/octet-stream", "text/plain", ""] or mime_type not in cls._parsers:
            if filename:
                import os
                _, ext = os.path.splitext(filename.lower())
                if ext in cls._extension_map:
                    resolved_mime = cls._extension_map[ext]
                    
        if resolved_mime not in cls._parsers:
            raise ValueError(f"Unsupported MIME type: {mime_type} (resolved: {resolved_mime})")
            
        return cls._parsers[resolved_mime]
